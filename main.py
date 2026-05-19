"""
Frequency-Weighted p-Median on Semantic Word Space
EMU414 - Mathematical Programming and Computer Applications
Hacettepe University, Spring 2025-2026

Model (Weighted p-Median):
    minimize    sum_{w in W} sum_{v in N(w)}  F_w * D_wv * y_wv

    s.t.
        sum_{v in N(w)}  y_wv  =  1           for all w in W   [assignment]
        y_wv  <=  x_v                          for all w,v      [linking]
        sum_{v not in known}  x_v   <=  B                       [budget]
        x_v  =  1                              for all v in known [known]
        x_v, y_wv  in  {0,1}                  for all w,v      [binary]

    N(w): if |W| > SparseThreshold, only the KNeighbors nearest words
          plus TopPopular most frequent words plus known words;
          otherwise all words.

Inputs:
    data/input.xlsx
        Sheet "Words"      : columns rank, word  (frequency-sorted)
        Sheet "Parameters" : columns Parameter, Value
                             rows: Budget, MIPGap, TimeLimit,
                                   SparseThreshold, KNeighbors, TopPopular
        Sheet "KnownWords" : single column of pre-learnt words (optional)
    data/dolma_300_2024_1.2M.100_combined.txt
        GloVe-format 300-dim embeddings

Outputs:
    model.lp              Gurobi LP export
    gurobi.log            Gurobi solver log
    results/solution.txt  human-readable solution
    <input>_solution.xlsx solution Excel
"""

import os
import sys
import numpy as np
import pandas as pd
import gurobipy as gp
from gurobipy import GRB

# ─────────────────────────────────────────────────────────────────────────────
# Paths
# ─────────────────────────────────────────────────────────────────────────────
EXCEL_FILE  = os.path.join("data", "input.xlsx")
EMBED_FILE  = os.path.join("data", "dolma_300_2024_1.2M.100_combined.txt")
RESULTS_DIR = "results"
LP_FILE     = "model.lp"
LOG_FILE    = "gurobi.log"
RESULT_FILE = os.path.join(RESULTS_DIR, "solution.txt")


# ─────────────────────────────────────────────────────────────────────────────
# Data loading
# ─────────────────────────────────────────────────────────────────────────────
def load_excel(path):
    if not os.path.exists(path):
        sys.exit(f"[ERROR] {path} not found.")

    words_df = pd.read_excel(path, sheet_name="Words")
    if not {"rank", "word"}.issubset(words_df.columns):
        sys.exit("[ERROR] Sheet 'Words' must have columns: rank, word")

    params_df = pd.read_excel(path, sheet_name="Parameters")
    if not {"Parameter", "Value"}.issubset(params_df.columns):
        sys.exit("[ERROR] Sheet 'Parameters' must have columns: Parameter, Value")

    params = dict(zip(params_df["Parameter"].str.strip(), params_df["Value"]))
    for key in ("Budget", "MIPGap", "TimeLimit"):
        if key not in params:
            sys.exit(f"[ERROR] Sheet 'Parameters' missing row: {key}")

    words    = words_df["word"].astype(str).tolist()
    ranks    = words_df["rank"].astype(float).tolist()
    budget   = int(params["Budget"])
    mipgap   = float(params["MIPGap"])
    tlimit   = float(params["TimeLimit"])
    sp_thresh = int(params.get("SparseThreshold", 2000))
    k_neighbors = int(params.get("KNeighbors", 200))
    top_popular = int(params.get("TopPopular", 10))

    weights = compute_freq_weights(ranks)

    xl = pd.ExcelFile(path)
    if "KnownWords" in xl.sheet_names:
        kw_df       = pd.read_excel(path, sheet_name="KnownWords")
        known_words = kw_df.iloc[:, 0].dropna().astype(str).tolist()
    else:
        known_words = []

    return words, weights, budget, mipgap, tlimit, known_words, sp_thresh, k_neighbors, top_popular


def compute_freq_weights(ranks):
    """Compute normalised frequency weights from rank list. Modify here to change weighting scheme."""
    raw   = [1.0 / r for r in ranks]
    max_w = max(raw)
    return [(w / max_w) for w in raw]


def load_embeddings(path, words):
    if not os.path.exists(path):
        sys.exit(
            f"[ERROR] {path} not found.\n"
            f"        Download embeddings from:\n"
            f"        https://nlp.stanford.edu/data/wordvecs/glove.2024.dolma.300d.zip\n"
            f"        Extract the .txt file into the data/ directory."
        )

    needed = set(words)
    found  = {}

    print(f"[INFO] Scanning embeddings for {len(needed)} words ...")
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            parts = line.split()
            token = parts[0]
            if token in needed:
                found[token] = np.array(parts[1:], dtype=np.float32)
                if len(found) == len(needed):
                    break

    missing = needed - set(found)
    if missing:
        print(f"[WARN] {len(missing)} words not found in embeddings: {missing}")

    dim = next(iter(found.values())).shape[0] if found else 0
    print(f"[INFO] Loaded {len(found)} embeddings (dim={dim}).")
    return found


def build_neighbor_structure(words, embeddings, known_indices, sp_thresh, k_neighbors, top_popular):
    """
    Returns V (n, dim), neighbors[w] (array of valid rep indices), dist_rows[w] (parallel distances).
    When n > sp_thresh, restricts each word's candidates to its k_neighbors nearest
    + top_popular most frequent + known_indices. Otherwise all words are candidates.
    """
    n    = len(words)
    zero = np.zeros(next(iter(embeddings.values())).shape[0], dtype=np.float32) if embeddings else np.zeros(300, dtype=np.float32)
    V    = np.stack([embeddings.get(w, zero) for w in words])

    if n <= sp_thresh:
        diff         = V[:, None, :] - V[None, :, :]
        dist_matrix  = np.sqrt((diff ** 2).sum(axis=-1))
        neighbors    = [np.arange(n, dtype=np.int32) for _ in range(n)]
        dist_rows    = [dist_matrix[w] for w in range(n)]
        print(f"[INFO] Full distance matrix ({n}x{n}).")
    else:
        always = set(range(min(top_popular, n))) | known_indices
        neighbors = [None] * n
        dist_rows  = [None] * n
        chunk = 512
        print(f"[INFO] Sparse mode: {k_neighbors} nearest + {top_popular} popular + {len(known_indices)} known per word.")
        for start in range(0, n, chunk):
            end   = min(start + chunk, n)
            block = V[start:end]                                           # (B, dim)
            dists = np.sqrt(((block[:, None, :] - V[None, :, :]) ** 2).sum(axis=2))  # (B, n)
            for i, w in enumerate(range(start, end)):
                nearest  = set(np.argsort(dists[i])[:k_neighbors])
                valid    = sorted(nearest | always)
                neighbors[w] = np.array(valid, dtype=np.int32)
                dist_rows[w] = dists[i, valid].astype(np.float64)
        print(f"[INFO] Sparse neighbor structure built.")

    return V, neighbors, dist_rows


# ─────────────────────────────────────────────────────────────────────────────
# Gurobi model  —  Weighted p-Median
# ─────────────────────────────────────────────────────────────────────────────
def build_and_solve(words, weights, neighbors, dist_rows, budget, mipgap, tlimit, known_indices):
    n = len(words)
    os.makedirs(RESULTS_DIR, exist_ok=True)

    model = gp.Model("pMedian_Words")
    model.setParam("LogFile",   LOG_FILE)
    model.setParam("MIPGap",    mipgap)
    model.setParam("TimeLimit", tlimit)

    x = model.addVars(n, vtype=GRB.BINARY, name="x")

    # Sparse y: only create variables for valid (w, v) pairs
    y = {}
    for w in range(n):
        for v in neighbors[w].tolist():
            y[w, v] = model.addVar(vtype=GRB.BINARY, name=f"y_{w}_{v}")
    model.update()

    # Force known words to be selected as representatives
    for v in known_indices:
        model.addConstr(x[v] == 1, name=f"known_{v}")

    model.setObjective(
        gp.quicksum(
            weights[w] * float(dist_rows[w][i]) * y[w, int(neighbors[w][i])]
            for w in range(n)
            for i in range(len(neighbors[w]))
        ),
        GRB.MINIMIZE
    )

    for w in range(n):
        model.addConstr(
            gp.quicksum(y[w, int(v)] for v in neighbors[w]) == 1,
            name=f"assign_{w}"
        )

    for w in range(n):
        for v in neighbors[w].tolist():
            model.addConstr(y[w, v] <= x[v], name=f"link_{w}_{v}")

    # Known words do not count toward budget
    free = [v for v in range(n) if v not in known_indices]
    model.addConstr(gp.quicksum(x[v] for v in free) <= budget, name="budget")

    model.write(LP_FILE)
    model.optimize()

    return model, x, y, n


# ─────────────────────────────────────────────────────────────────────────────
# Results
# ─────────────────────────────────────────────────────────────────────────────
def write_results(model, words, weights, V, neighbors, x, y, n, budget, input_file):
    if model.Status not in (GRB.OPTIMAL, GRB.TIME_LIMIT):
        print(f"[WARN] Solver status: {model.Status}. No feasible solution.")
        return

    status_str = "OPTIMAL" if model.Status == GRB.OPTIMAL else "TIME LIMIT"

    reps = [v for v in range(n) if x[v].X > 0.5]

    assignments = {}
    for w in range(n):
        for v in neighbors[w].tolist():
            if y[w, v].X > 0.5:
                assignments[w] = v
                break

    total_weighted_dist = model.ObjVal

    summary = [
        "=" * 65,
        "  FREQUENCY-WEIGHTED p-MEDIAN — SOLUTION",
        "=" * 65,
        f"  Status                       : {status_str}",
        f"  Objective (min weighted dist) : {total_weighted_dist:.6f}",
        f"  Budget (B)                    : {budget}",
        "=" * 65,
        "",
        f"SELECTED REPRESENTATIVES ({len(reps)}):",
    ]
    for v in reps:
        summary.append(f"  [{v:3d}]  {words[v]:<25}  (freq_weight = {weights[v]:.4f})")

    assignment_lines = ["", "ASSIGNMENTS  (word  ->  representative  |  L2 distance):"]
    for w in range(n):
        v      = assignments.get(w, -1)
        rep    = words[v] if v >= 0 else "???"
        d      = float(np.linalg.norm(V[w] - V[v])) if v >= 0 else -1.0
        marker = " (self)" if w == v else ""
        assignment_lines.append(
            f"  {words[w]:<25}  ->  {rep:<25}  d={d:.4f}{marker}"
        )

    os.makedirs(RESULTS_DIR, exist_ok=True)
    with open(RESULT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(summary + assignment_lines))

    print("\n".join(summary))

    solution_path = write_solution_excel(
        model, words, weights, V, reps, assignments,
        n, budget, status_str, total_weighted_dist, input_file
    )

    print(f"\n[INFO] Solution Excel  : {solution_path}")
    print(f"[INFO] Results txt     : {RESULT_FILE}")
    print(f"[INFO] LP file         : {LP_FILE}")
    print(f"[INFO] Gurobi log      : {LOG_FILE}")


def write_solution_excel(model, words, weights, V, reps, assignments,
                         n, budget, status_str, obj_val, input_file):
    from openpyxl import load_workbook

    out_path = os.path.splitext(input_file)[0] + "_solution.xlsx"
    wb = load_workbook(input_file)

    # ── Sheet: Representatives ───────────────────────────────────────────────
    ws_reps = wb.create_sheet("Representatives")
    ws_reps.append(["Index", "Word", "Freq_Weight"])
    for v in reps:
        ws_reps.append([v, words[v], round(weights[v], 6)])

    # ── Sheet: Assignments ───────────────────────────────────────────────────
    ws_asgn = wb.create_sheet("Assignments")
    ws_asgn.append(["Word", "Freq_Weight", "Representative", "L2_Distance", "Is_Self"])
    for w in range(n):
        v   = assignments.get(w, -1)
        rep = words[v] if v >= 0 else ""
        d   = round(float(np.linalg.norm(V[w] - V[v])), 6) if v >= 0 else ""
        ws_asgn.append([words[w], round(weights[w], 6), rep, d, w == v])

    # ── Sheet: Solver ────────────────────────────────────────────────────────
    ws_solv = wb.create_sheet("Solver")
    ws_solv.append(["Parameter", "Value"])
    ws_solv.append(["Status",      status_str])
    ws_solv.append(["Objective",   round(obj_val, 8)])
    ws_solv.append(["SolveTime_s", round(model.Runtime, 4)])
    ws_solv.append(["MIPGap",      round(model.MIPGap, 8)])
    ws_solv.append(["NodeCount",   int(model.NodeCount)])
    ws_solv.append(["SolCount",    int(model.SolCount)])
    ws_solv.append(["Budget",      budget])

    wb.save(out_path)
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    args       = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags      = {a for a in sys.argv[1:] if a.startswith("--")}
    input_file = args[0] if args else EXCEL_FILE
    uniform    = "--uniform" in flags

    print("[INFO] Loading Excel ...")
    words, weights, budget, mipgap, tlimit, known_words, sp_thresh, k_neighbors, top_popular = load_excel(input_file)
    if uniform:
        weights = [1.0] * len(words)
        print("[INFO] --uniform flag set: all freq_weights = 1")

    word_index    = {w: i for i, w in enumerate(words)}
    known_indices = set()
    for kw in known_words:
        if kw in word_index:
            known_indices.add(word_index[kw])
        else:
            print(f"[WARN] Known word '{kw}' not in Words sheet — skipped.")

    print(f"[INFO] {len(words)} words | B={budget} | known={len(known_indices)} | MIPGap={mipgap} | TimeLimit={tlimit}s")

    embeddings = load_embeddings(EMBED_FILE, words)
    V, neighbors, dist_rows = build_neighbor_structure(
        words, embeddings, known_indices, sp_thresh, k_neighbors, top_popular
    )

    print(f"[INFO] Solving p-Median (B={budget}) ...")
    model, x, y, n = build_and_solve(
        words, weights, neighbors, dist_rows, budget, mipgap, tlimit, known_indices
    )
    write_results(model, words, weights, V, neighbors, x, y, n, budget, input_file)
